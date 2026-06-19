from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Count, Avg, Max, Case, When, Value, F, FloatField
from django.db.models.functions import Coalesce, TruncMonth
from django.core.paginator import Paginator
from functools import wraps
from datetime import datetime, timedelta
import json
import csv
from io import BytesIO
from django.http import HttpResponse, JsonResponse
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .models import Student, Company, Job, Application, Interview, PlacementResult, Notification
from .forms import (
    StudentRegistrationForm, StudentLoginForm, StudentPasswordChangeForm,
    StudentProfileForm, CompanyForm, JobForm, ApplicationStatusForm, InterviewForm
)

# ============================================================================
# ADMIN ACCESS DECORATOR
# ============================================================================

def admin_required(view_func):
    """
    Decorator to check if user is admin (staff).
    Redirects non-admin users to home page.
    """
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not request.user.is_staff:
            messages.error(request, 'You do not have permission to access the admin panel.')
            return redirect('home')
        return view_func(request, *args, **kwargs)
    return wrapper

# ============================================================================
# HOME VIEW
# ============================================================================

def home(request):
    """
    Home page of the Placement Management System.
    Redirects admin to admin dashboard and students to student dashboard if already logged in.
    """
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('admin_dashboard')
        else:
            try:
                request.user.student_profile
                return redirect('student_dashboard')
            except Student.DoesNotExist:
                return redirect('student_profile_create')
    
    context = {
        'total_companies': Company.objects.count(),
        'total_jobs': Job.objects.count(),
        'total_students': Student.objects.count(),
    }
    return render(request, 'home.html', context)


# ============================================================================
# AUTHENTICATION VIEWS
# ============================================================================

@require_http_methods(["GET", "POST"])
def student_register(request):
    """
    Register a new student.
    GET: Display registration form
    POST: Process registration and create student profile
    """
    if request.method == 'POST':
        form = StudentRegistrationForm(request.POST)
        if form.is_valid():
            # Create user
            user = form.save()
            # Automatically login the user after registration
            login(request, user)
            messages.success(request, 'Registration successful! Please complete your profile.')
            # Redirect to profile creation
            return redirect('student_profile_create')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = StudentRegistrationForm()
    
    return render(request, 'student/register.html', {'form': form})


@require_http_methods(["GET", "POST"])
def student_login(request):
    """
    Login page for students.
    GET: Display login form
    POST: Authenticate and login student
    """
    if request.method == 'POST':
        form = StudentLoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Welcome back, {user.get_full_name()}!')
                # Check if user is admin
                if user.is_staff:
                    return redirect('admin_dashboard')
                else:
                    return redirect('student_dashboard')
            else:
                messages.error(request, 'Invalid email or password.')
        else:
            messages.error(request, 'Invalid email or password.')
    else:
        form = StudentLoginForm()
    
    return render(request, 'student/login.html', {'form': form})


@login_required(login_url='login')
def student_logout(request):
    """
    Logout the current student.
    """
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('home')


@login_required(login_url='login')
@require_http_methods(["GET", "POST"])
def student_change_password(request):
    """
    Change password for logged-in student.
    """
    if request.method == 'POST':
        form = StudentPasswordChangeForm(request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Password changed successfully!')
            return redirect('student_dashboard')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = StudentPasswordChangeForm(request.user)
    
    return render(request, 'student/change_password.html', {'form': form})


# ============================================================================
# STUDENT PROFILE VIEWS
# ============================================================================

@login_required(login_url='login')
@require_http_methods(["GET", "POST"])
def student_profile_create(request):
    """
    Create student profile for newly registered students.
    """
    # Check if student already has a profile
    try:
        student = request.user.student_profile
        return redirect('student_profile')
    except Student.DoesNotExist:
        pass
    
    if request.method == 'POST':
        form = StudentProfileForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save(commit=False)
            student.user = request.user
            student.save()
            messages.success(request, 'Profile created successfully!')
            return redirect('student_profile')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = StudentProfileForm()
    
    return render(request, 'student/profile_create.html', {'form': form})


@login_required(login_url='login')
@require_http_methods(["GET", "POST"])
def student_profile(request):
    """
    View and edit student profile.
    """
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.info(request, 'Please complete your profile first.')
        return redirect('student_profile_create')
    
    if request.method == 'POST':
        form = StudentProfileForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('student_profile')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = StudentProfileForm(instance=student)
    
    # Process skills - split comma-separated string into list
    skills_list = []
    if student.skills:
        skills_list = [skill.strip() for skill in student.skills.split(',') if skill.strip()]
    
    context = {
        'form': form,
        'student': student,
        'skills_list': skills_list
    }
    return render(request, 'student/profile.html', context)


# ============================================================================
# STUDENT DASHBOARD
# ============================================================================

@login_required(login_url='login')
def student_dashboard(request):
    """
    Student dashboard showing applications and interview status.
    """
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        return redirect('student_profile_create')
    
    # Get all applications for this student
    applications = student.applications.all()
    
    # Calculate statistics
    stats = {
        'total_applications': applications.count(),
        'shortlisted': applications.filter(status='Shortlisted').count(),
        'interview_scheduled': applications.filter(status='Interview Scheduled').count(),
        'selected': applications.filter(status='Selected').count(),
        'rejected': applications.filter(status='Rejected').count(),
    }
    
    context = {
        'student': student,
        'applications': applications,
        'stats': stats,
    }
    return render(request, 'student/dashboard.html', context)


# ============================================================================
# JOB VIEWS
# ============================================================================

@login_required(login_url='login')
def job_list(request):
    """
    List all available jobs for students.
    Shows eligibility status based on CGPA.
    """
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.info(request, 'Please complete your profile to view jobs.')
        return redirect('student_profile_create')
    
    # Get all active jobs
    jobs = Job.objects.filter(deadline__gte=__import__('datetime').datetime.now().date()).order_by('-deadline')
    
    # Add eligibility status for each job
    for job in jobs:
        job.is_eligible = student.cgpa >= job.min_cgpa
        # Check if student has already applied
        job.already_applied = Application.objects.filter(
            student=student,
            job=job
        ).exists()
    
    # Pagination
    paginator = Paginator(jobs, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'jobs': page_obj.object_list,
        'student': student,
    }
    return render(request, 'student/jobs.html', context)


@login_required(login_url='login')
def job_detail(request, job_id):
    """
    View detailed information about a specific job.
    """
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.info(request, 'Please complete your profile to view job details.')
        return redirect('student_profile_create')
    
    job = get_object_or_404(Job, id=job_id)
    
    # Check eligibility and if already applied
    is_eligible = student.cgpa >= job.min_cgpa
    already_applied = Application.objects.filter(student=student, job=job).exists()
    
    context = {
        'job': job,
        'student': student,
        'is_eligible': is_eligible,
        'already_applied': already_applied,
    }
    return render(request, 'student/job_detail.html', context)


# ============================================================================
# APPLICATION VIEWS
# ============================================================================

@login_required(login_url='login')
@require_http_methods(["POST"])
def apply_for_job(request, job_id):
    """
    Apply for a specific job (POST only).
    """
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.error(request, 'Please complete your profile first.')
        return redirect('student_profile_create')
    
    job = get_object_or_404(Job, id=job_id)
    
    # Check if eligible
    if student.cgpa < job.min_cgpa:
        messages.error(request, f'Your CGPA ({student.cgpa}) is below the minimum requirement ({job.min_cgpa}) for this job.')
        return redirect('job_detail', job_id=job_id)
    
    # Check if already applied
    if Application.objects.filter(student=student, job=job).exists():
        messages.info(request, 'You have already applied for this job.')
        return redirect('job_detail', job_id=job_id)
    
    # Create application
    application = Application.objects.create(student=student, job=job)
    
    # Generate notification for application submission
    create_notification(
        student=student,
        notification_type='application_submitted',
        title='Application Submitted',
        message=f"Your application for {job.job_title} at {job.company.company_name} has been submitted successfully.",
        application=application,
        job=job
    )
    
    messages.success(request, f'You have successfully applied for {job.job_title} at {job.company.company_name}.')
    return redirect('student_applications')


@login_required(login_url='login')
def student_applications(request):
    """
    View all applications made by the student.
    """
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        return redirect('student_profile_create')
    
    applications = student.applications.all().order_by('-applied_date')
    
    # Pagination
    paginator = Paginator(applications, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'applications': page_obj.object_list,
    }
    return render(request, 'student/applications.html', context)


# ============================================================================
# ADMIN DASHBOARD
# ============================================================================

@admin_required
def admin_dashboard(request):
    """
    Admin dashboard showing comprehensive placement statistics and KPIs.
    Displays professional KPI cards, analytics, and recent activity.
    Updated to include PlacementResult metrics.
    """
    # ========== KPI ROW 1: Core Student Metrics ==========
    total_students = Student.objects.count()
    # Get placed students from PlacementResult (Selected status) + Application (Selected status for backward compatibility)
    placed_by_result = PlacementResult.objects.filter(status='Selected').values('student').distinct().count()
    placed_by_app = Application.objects.filter(status='Selected').exclude(
        id__in=PlacementResult.objects.values_list('application_id', flat=True)
    ).values('student').distinct().count()
    placed_students_count = placed_by_result + placed_by_app
    
    eligible_students = total_students  # Could be refined based on CGPA cutoff
    placement_percentage = (placed_students_count / total_students * 100) if total_students > 0 else 0
    
    # ========== KPI ROW 2: Drive & Interview Metrics ==========
    total_companies = Company.objects.count()
    active_drives = Job.objects.filter(deadline__gte=datetime.now().date()).count()
    interviews_scheduled = Interview.objects.count()
    pending_applications = Application.objects.filter(status='Applied').count()
    
    # ========== KPI ROW 3: Package & Selection Metrics (Using PlacementResult) ==========
    # Get packages from PlacementResult
    placement_results = PlacementResult.objects.filter(status='Selected')
    packages = [float(result.package_offered) for result in placement_results if result.package_offered]
    
    highest_package = max(packages) if packages else 0
    average_package = sum(packages) / len(packages) if packages else 0
    total_offers = placement_results.count()
    total_selections = placement_results.values('student').distinct().count()
    
    # ========== Recent Activity ==========
    recent_activities = []
    
    # Get recent applications (last 5)
    recent_apps = Application.objects.select_related('student', 'job__company').order_by('-applied_date')[:3]
    for app in recent_apps:
        recent_activities.append({
            'type': 'application',
            'text': f"{app.student.user.get_full_name()} applied to {app.job.job_title}",
            'timestamp': app.applied_date,
            'icon': 'fa-file-alt',
            'color': 'info'
        })
    
    # Get recent job postings (last 5)
    recent_jobs = Job.objects.select_related('company').order_by('-created_at')[:3]
    for job in recent_jobs:
        recent_activities.append({
            'type': 'job',
            'text': f"{job.company.company_name} created drive for {job.job_title}",
            'timestamp': job.created_at,
            'icon': 'fa-briefcase',
            'color': 'success'
        })
    
    # Get recent interviews (last 5)
    recent_interviews = Interview.objects.select_related('application__student', 'application__job__company').order_by('-created_at')[:3]
    for interview in recent_interviews:
        recent_activities.append({
            'type': 'interview',
            'text': f"Interview scheduled for {interview.application.student.user.get_full_name()} at {interview.application.job.company.company_name}",
            'timestamp': interview.created_at,
            'icon': 'fa-calendar-alt',
            'color': 'warning'
        })
    
    # Get recent placement results
    recent_results = PlacementResult.objects.select_related('student', 'company').order_by('-published_date')[:3]
    for result in recent_results:
        status_icon = 'fa-check-circle' if result.status == 'Selected' else 'fa-times-circle'
        status_color = 'success' if result.status == 'Selected' else 'danger'
        recent_activities.append({
            'type': 'result',
            'text': f"{result.student.user.get_full_name()} - {result.status} by {result.company.company_name}",
            'timestamp': result.published_date,
            'icon': status_icon,
            'color': status_color
        })
    
    # Sort recent activities by timestamp (newest first)
    recent_activities = sorted(recent_activities, key=lambda x: x['timestamp'], reverse=True)[:10]
    
    # ========== Upcoming Interviews (Next 7 days) ==========
    today = datetime.now().date()
    next_week = today + timedelta(days=7)
    upcoming_interviews = Interview.objects.select_related(
        'application__student', 
        'application__job__company'
    ).filter(
        date__gte=today,
        date__lte=next_week
    ).order_by('date', 'time')[:8]
    
    # ========== Analytics Data for Charts ==========
    
    # 1. Students Placed per Branch
    students_per_branch = Student.objects.values('branch').annotate(
        total=Count('id'),
        placed=Count(Case(When(placement_results__status='Selected', then=1)))
    ).order_by('branch')
    
    branch_names = [item['branch'] for item in students_per_branch]
    branch_totals = [item['total'] for item in students_per_branch]
    branch_placed = [item['placed'] for item in students_per_branch]
    
    # 2. Company-wise Hiring Count
    company_hiring = PlacementResult.objects.filter(
        status='Selected'
    ).values('company__company_name').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    company_names = [item['company__company_name'] for item in company_hiring]
    company_counts = [item['count'] for item in company_hiring]
    
    # 3. Monthly Placements (from PlacementResult)
    monthly_placements = PlacementResult.objects.annotate(
        month=TruncMonth('selection_date')
    ).values('month').annotate(
        count=Count('id')
    ).filter(status='Selected').order_by('month')
    
    # Format dates for display
    months = []
    placement_counts = []
    for item in monthly_placements:
        if item['month']:
            months.append(item['month'].strftime('%b %Y'))
            placement_counts.append(item['count'])
    
    # 4. Application Status Distribution
    status_distribution = Application.objects.values('status').annotate(
        count=Count('id')
    ).order_by('status')
    
    status_labels = [item['status'] for item in status_distribution]
    status_counts = [item['count'] for item in status_distribution]
    
    # Convert numeric data to JSON for Chart.js
    context = {
        # KPI Row 1
        'total_students': total_students,
        'eligible_students': eligible_students,
        'placed_students': placed_students_count,
        'placement_percentage': round(placement_percentage, 2),
        
        # KPI Row 2
        'total_companies': total_companies,
        'active_drives': active_drives,
        'interviews_scheduled': interviews_scheduled,
        'pending_applications': pending_applications,
        
        # KPI Row 3
        'highest_package': round(float(highest_package), 2),
        'average_package': round(float(average_package), 2),
        'total_offers': total_offers,
        'total_selections': total_selections,
        
        # Recent Activity
        'recent_activities': recent_activities,
        
        # Upcoming Interviews
        'upcoming_interviews': upcoming_interviews,
        
        # Chart Data (as JSON strings)
        'branch_names': json.dumps(branch_names),
        'branch_totals': json.dumps(branch_totals),
        'branch_placed': json.dumps(branch_placed),
        'company_names': json.dumps(company_names),
        'company_counts': json.dumps(company_counts),
        'months': json.dumps(months),
        'placement_counts': json.dumps(placement_counts),
        'status_labels': json.dumps(status_labels),
        'status_counts': json.dumps(status_counts),
    }
    return render(request, 'admin_panel/dashboard.html', context)


# ============================================================================
# ADMIN COMPANY MANAGEMENT
# ============================================================================

@login_required(login_url='login')
def admin_company_list(request):
    """
    Admin view to list all companies.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    companies = Company.objects.all().order_by('company_name')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        companies = companies.filter(
            Q(company_name__icontains=search_query) |
            Q(hr_name__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(companies, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'companies': page_obj.object_list,
        'search_query': search_query,
    }
    return render(request, 'admin_panel/company_list.html', context)


@admin_required
@require_http_methods(["GET", "POST"])
def admin_company_create(request):
    """
    Admin view to create a new company.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    if request.method == 'POST':
        form = CompanyForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Company added successfully!')
            return redirect('admin_company_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = CompanyForm()
    
    return render(request, 'admin_panel/company_form.html', {'form': form, 'is_create': True})


@admin_required
@require_http_methods(["GET", "POST"])
def admin_company_edit(request, company_id):
    """
    Admin view to edit a company.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    company = get_object_or_404(Company, id=company_id)
    
    if request.method == 'POST':
        form = CompanyForm(request.POST, instance=company)
        if form.is_valid():
            form.save()
            messages.success(request, 'Company updated successfully!')
            return redirect('admin_company_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = CompanyForm(instance=company)
    
    return render(request, 'admin_panel/company_form.html', {'form': form, 'company': company})


@admin_required
@require_http_methods(["POST"])
def admin_company_delete(request, company_id):
    """
    Admin view to delete a company.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    company = get_object_or_404(Company, id=company_id)
    company_name = company.company_name
    company.delete()
    messages.success(request, f'Company "{company_name}" deleted successfully!')
    return redirect('admin_company_list')


# ============================================================================
# ADMIN JOB MANAGEMENT
# ============================================================================

@login_required(login_url='login')
def admin_job_list(request):
    """
    Admin view to list all jobs.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    jobs = Job.objects.all().order_by('-deadline')
    
    # Search and filter
    search_query = request.GET.get('search', '')
    if search_query:
        jobs = jobs.filter(
            Q(job_title__icontains=search_query) |
            Q(company__company_name__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(jobs, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'jobs': page_obj.object_list,
        'search_query': search_query,
    }
    return render(request, 'admin_panel/job_list.html', context)


@admin_required
@require_http_methods(["GET", "POST"])
def admin_job_create(request):
    """
    Admin view to create a new job.
    Creates notifications for all students when job is posted.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    if request.method == 'POST':
        form = JobForm(request.POST)
        if form.is_valid():
            job = form.save()
            
            # Create notification for all students
            all_students = Student.objects.all()
            company = job.company
            for student in all_students:
                create_notification(
                    student=student,
                    notification_type='job_posted',
                    title='New Job Opportunity',
                    message=f"A new job has been posted by {company.company_name}. Package: {company.package} LPA.",
                    job=job
                )
            
            messages.success(request, 'Job posted successfully!')
            return redirect('admin_job_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = JobForm()
    
    return render(request, 'admin_panel/job_form.html', {'form': form, 'is_create': True})


@admin_required
@require_http_methods(["GET", "POST"])
def admin_job_edit(request, job_id):
    """
    Admin view to edit a job.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    job = get_object_or_404(Job, id=job_id)
    
    if request.method == 'POST':
        form = JobForm(request.POST, instance=job)
        if form.is_valid():
            form.save()
            messages.success(request, 'Job updated successfully!')
            return redirect('admin_job_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = JobForm(instance=job)
    
    return render(request, 'admin_panel/job_form.html', {'form': form, 'job': job})


@admin_required
@require_http_methods(["POST"])
def admin_job_delete(request, job_id):
    """
    Admin view to delete a job.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    job = get_object_or_404(Job, id=job_id)
    job_title = job.job_title
    job.delete()
    messages.success(request, f'Job "{job_title}" deleted successfully!')
    return redirect('admin_job_list')


# ============================================================================
# ADMIN APPLICATION MANAGEMENT
# ============================================================================

@login_required(login_url='login')
def admin_applications(request):
    """
    Admin view to manage all applications.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    applications = Application.objects.all().order_by('-applied_date')
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        applications = applications.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(applications, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'applications': page_obj.object_list,
        'status_filter': status_filter,
    }
    return render(request, 'admin_panel/applications.html', context)


@admin_required
@require_http_methods(["GET", "POST"])
@admin_required
@require_http_methods(["GET", "POST"])
def admin_application_update(request, application_id):
    """
    Admin view to update application status.
    Status changes automatically trigger all related business logic:
    - Generates appropriate notifications
    - Creates PlacementResult records for Selected/Rejected status
    - Maintains data consistency across the system
    
    Application Status is the single source of truth.
    No separate "Publish Result" workflow needed.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    application = get_object_or_404(Application, id=application_id)
    old_status = application.status  # Capture old status before any changes
    
    if request.method == 'POST':
        form = ApplicationStatusForm(request.POST, instance=application)
        if form.is_valid():
            # Save the form first to update application status in database
            updated_application = form.save(commit=True)
            new_status = updated_application.status
            
            # Only proceed if status actually changed
            if old_status != new_status:
                # Let the helper function handle all status transition logic
                # (notifications, PlacementResult creation, etc.)
                handle_application_status_transition(updated_application, new_status, old_status, request.user)
                messages.success(request, f'✓ Application status updated from "{old_status}" to "{new_status}"')
            else:
                messages.info(request, 'No changes made - status is the same.')
            
            return redirect('admin_applications')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = ApplicationStatusForm(instance=application)
    
    context = {
        'form': form,
        'application': application,
    }
    return render(request, 'admin_panel/application_update.html', context)


# ============================================================================
# ADMIN INTERVIEW MANAGEMENT
# ============================================================================

@admin_required
def admin_interviews(request):
    """
    Admin view to list all scheduled interviews.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    interviews = Interview.objects.all().order_by('date', 'time')
    
    # Pagination
    paginator = Paginator(interviews, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'interviews': page_obj.object_list,
    }
    return render(request, 'admin_panel/interviews.html', context)


@admin_required
@require_http_methods(["GET", "POST"])
def admin_interview_schedule(request, application_id):
    """
    Admin view to schedule an interview for an application.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    application = get_object_or_404(Application, id=application_id)
    
    # Check if interview already scheduled
    if hasattr(application, 'interview'):
        messages.info(request, 'Interview already scheduled for this application.')
        return redirect('admin_interviews')
    
    if request.method == 'POST':
        form = InterviewForm(request.POST)
        if form.is_valid():
            interview = form.save(commit=False)
            interview.application = application
            interview.save()
            # Update application status
            application.status = 'Interview Scheduled'
            application.save()
            
            # Create notification for student
            interview_datetime = f"{interview.date.strftime('%d %B %Y')} at {interview.time.strftime('%H:%M')}"
            create_notification(
                student=application.student,
                notification_type='interview_scheduled',
                title='Interview Scheduled',
                message=f"Your interview for {application.job.job_title} at {application.job.company.company_name} "
                        f"has been scheduled on {interview_datetime}.",
                application=application
            )
            
            messages.success(request, 'Interview scheduled successfully!')
            return redirect('admin_interviews')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = InterviewForm()
    
    context = {
        'form': form,
        'application': application,
    }
    return render(request, 'admin_panel/interview_form.html', context)


@admin_required
@require_http_methods(["GET", "POST"])
def admin_interview_edit(request, interview_id):
    """
    Admin view to edit scheduled interview.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    interview = get_object_or_404(Interview, id=interview_id)
    
    if request.method == 'POST':
        form = InterviewForm(request.POST, instance=interview)
        if form.is_valid():
            form.save()
            messages.success(request, 'Interview updated successfully!')
            return redirect('admin_interviews')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = InterviewForm(instance=interview)
    
    context = {
        'form': form,
        'interview': interview,
    }
    return render(request, 'admin_panel/interview_form.html', context)


# ============================================================================
# ADMIN STUDENT MANAGEMENT
# ============================================================================

@admin_required
def admin_students(request):
    """
    Admin view to list all students.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    students = Student.objects.all().order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        students = students.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(roll_number__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(students, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'students': page_obj.object_list,
        'search_query': search_query,
    }
    return render(request, 'admin_panel/students.html', context)


@admin_required
def admin_student_detail(request, student_id):
    """
    Admin view to see detailed information about a student.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    student = get_object_or_404(Student, id=student_id)
    applications = student.applications.all().order_by('-applied_date')
    
    # Process skills from comma-separated string to list
    skills_list = []
    if student.skills:
        skills_list = [skill.strip() for skill in student.skills.split(',')]
    
    context = {
        'student': student,
        'applications': applications,
        'skills_list': skills_list,
    }
    return render(request, 'admin_panel/student_detail.html', context)


@admin_required
@require_http_methods(["GET", "POST"])
def admin_student_edit(request, student_id):
    """
    Admin view to edit student details.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == 'POST':
        form = StudentProfileForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, f'Student {student.user.get_full_name()} updated successfully!')
            return redirect('admin_student_detail', student_id=student_id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = StudentProfileForm(instance=student)
    
    context = {
        'form': form,
        'student': student,
    }
    return render(request, 'admin_panel/student_form.html', context)


@admin_required
@require_http_methods(["POST"])
def admin_student_delete(request, student_id):
    """
    Admin view to delete a student.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    student = get_object_or_404(Student, id=student_id)
    user = student.user
    student_name = student.user.get_full_name()
    
    # Delete student profile first, then the associated user
    student.delete()
    user.delete()
    
    messages.success(request, f'Student "{student_name}" and associated account deleted successfully!')
    return redirect('admin_students')


# ============================================================================
# PLACEMENT REPORTS MODULE
# ============================================================================

# Export Utility Functions
def export_to_csv(filename, headers, rows):
    """Generate CSV export"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerows(rows)
    return response

def export_to_excel(filename, headers, rows, sheet_name='Report'):
    """Generate Excel export"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Add headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response

@admin_required
def admin_reports(request):
    """
    Main reports dashboard with multiple report categories.
    """
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('home')
    
    # Analytics data for charts
    total_students = Student.objects.count()
    total_companies = Company.objects.count()
    total_jobs = Job.objects.count()
    total_applications = Application.objects.count()
    
    # Placement statistics
    placed_students = Application.objects.filter(status='Selected').values('student').distinct().count()
    placed_percentage = round((placed_students / total_students * 100) if total_students > 0 else 0, 2)
    
    # Branch-wise placements
    branch_stats = Student.objects.values('branch').annotate(
        total=Count('id'),
        placed=Count('applications', filter=Q(applications__status='Selected'))
    ).order_by('branch')
    
    # Company-wise hiring
    company_hiring = Company.objects.annotate(
        total_applicants=Count('jobs__applications'),
        selected_count=Count('jobs__applications', filter=Q(jobs__applications__status='Selected'))
    ).order_by('-total_applicants')[:10]
    
    # Application status distribution
    status_dist = {}
    for status in ['Applied', 'Shortlisted', 'Interview Scheduled', 'Selected', 'Rejected']:
        status_dist[status] = Application.objects.filter(status=status).count()
    
    # Package analysis
    selected_apps = Application.objects.filter(status='Selected').select_related('job__company')
    packages = [app.job.company.package for app in selected_apps]
    max_package = max(packages) if packages else 0
    min_package = min(packages) if packages else 0
    avg_package = sum(packages) / len(packages) if packages else 0
    
    # Monthly trends
    monthly_apps = Application.objects.annotate(month=TruncMonth('applied_date')).values('month').annotate(
        count=Count('id'), selected=Count('id', filter=Q(status='Selected'))
    ).order_by('month')
    
    # Convert monthly data to JSON-serializable format
    monthly_data_list = []
    for item in monthly_apps:
        monthly_data_list.append({
            'month': item['month'].strftime('%Y-%m') if item['month'] else '',
            'count': item['count'],
            'selected': item['selected']
        })
    
    # Prepare chart data
    chart_data = {
        'branch_names': [b['branch'] for b in branch_stats],
        'branch_total': [b['total'] for b in branch_stats],
        'branch_placed': [b['placed'] for b in branch_stats],
        'company_names': [c.company_name[:20] for c in company_hiring],
        'company_applicants': [c.total_applicants for c in company_hiring],
        'company_selected': [c.selected_count for c in company_hiring],
        'status_labels': list(status_dist.keys()),
        'status_values': list(status_dist.values()),
        'monthly_data': monthly_data_list,
    }
    
    context = {
        'total_students': total_students,
        'total_companies': total_companies,
        'total_jobs': total_jobs,
        'total_applications': total_applications,
        'placed_students': placed_students,
        'placed_percentage': placed_percentage,
        'max_package': max_package,
        'min_package': min_package,
        'avg_package': round(avg_package, 2),
        'chart_data': json.dumps(chart_data),
        'branch_stats': branch_stats,
        'company_hiring': company_hiring,
        'status_dist': status_dist,
    }
    return render(request, 'admin_panel/reports.html', context)

@admin_required
def student_report(request):
    """
    Detailed student report with search, filter, and export.
    """
    students = Student.objects.select_related('user').all()
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        students = students.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(roll_number__icontains=search_query)
        )
    
    # Filter by branch
    branch_filter = request.GET.get('branch', '')
    if branch_filter:
        students = students.filter(branch=branch_filter)
    
    # Filter by CGPA
    cgpa_filter = request.GET.get('cgpa', '')
    if cgpa_filter:
        cgpa_val = float(cgpa_filter)
        students = students.filter(cgpa__gte=cgpa_val)
    
    # Filter by graduation year
    year_filter = request.GET.get('graduation_year', '')
    if year_filter:
        students = students.filter(graduation_year=int(year_filter))
    
    # Build report data
    report_data = []
    for student in students:
        placements = Application.objects.filter(
            student=student, status='Selected'
        ).select_related('job__company').first()
        
        report_data.append({
            'student': student,
            'placed': placements is not None,
            'company': placements.job.company.company_name if placements else 'Not Placed',
            'package': placements.job.company.package if placements else 0,
        })
    
    # Export functionality
    export_format = request.GET.get('export', '')
    if export_format:
        headers = ['Name', 'Roll Number', 'Branch', 'CGPA', 'Graduation Year', 'Placed', 'Company', 'Package (LPA)']
        rows = []
        for data in report_data:
            rows.append([
                data['student'].user.get_full_name(),
                data['student'].roll_number,
                data['student'].branch,
                float(data['student'].cgpa),
                data['student'].graduation_year,
                'Yes' if data['placed'] else 'No',
                data['company'],
                float(data['package']) if data['package'] else 0,
            ])
        
        if export_format == 'csv':
            return export_to_csv('student_report', headers, rows)
        elif export_format == 'excel':
            return export_to_excel('student_report', headers, rows, 'Students')
    
    # Pagination
    paginator = Paginator(report_data, 25)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    
    branches = Student.BRANCH_CHOICES
    graduation_years = sorted(set(Student.objects.values_list('graduation_year', flat=True)))
    
    context = {
        'page_obj': page_obj,
        'report_data': page_obj,
        'search_query': search_query,
        'branch_filter': branch_filter,
        'cgpa_filter': cgpa_filter,
        'year_filter': year_filter,
        'branches': branches,
        'graduation_years': graduation_years,
    }
    return render(request, 'admin_panel/reports/student_report.html', context)

@admin_required
def company_report(request):
    """
    Company-wise placement analysis report.
    """
    companies = Company.objects.annotate(
        total_applicants=Count('jobs__applications'),
        shortlisted=Count('jobs__applications', filter=Q(jobs__applications__status='Shortlisted')),
        selected=Count('jobs__applications', filter=Q(jobs__applications__status='Selected')),
        interviewed=Count('jobs__applications', filter=Q(jobs__applications__status='Interview Scheduled')),
        rejected=Count('jobs__applications', filter=Q(jobs__applications__status='Rejected')),
    ).order_by('-total_applicants')
    
    # Search by company name
    search_query = request.GET.get('search', '')
    if search_query:
        companies = companies.filter(company_name__icontains=search_query)
    
    # Sort functionality
    sort_by = request.GET.get('sort', '-total_applicants')
    companies = companies.order_by(sort_by)
    
    # Export functionality
    export_format = request.GET.get('export', '')
    if export_format:
        headers = ['Company Name', 'Total Applicants', 'Shortlisted', 'Interviewed', 'Selected', 'Rejected', 'Package (LPA)']
        rows = []
        for company in companies:
            rows.append([
                company.company_name,
                company.total_applicants,
                company.shortlisted,
                company.interviewed,
                company.selected,
                company.rejected,
                float(company.package),
            ])
        
        if export_format == 'csv':
            return export_to_csv('company_report', headers, rows)
        elif export_format == 'excel':
            return export_to_excel('company_report', headers, rows, 'Companies')
    
    # Pagination
    paginator = Paginator(companies, 15)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'sort_by': sort_by,
    }
    return render(request, 'admin_panel/reports/company_report.html', context)

@admin_required
def placement_report(request):
    """
    Comprehensive placement analysis with multiple views.
    """
    report_type = request.GET.get('type', 'student')
    
    if report_type == 'branch':
        # Branch-wise placement report
        branches = Student.objects.values('branch').annotate(
            total_students=Count('id'),
            placed=Count('applications', filter=Q(applications__status='Selected')),
            avg_cgpa=Avg('cgpa'),
        ).order_by('branch')
        
        # Calculate package stats per branch
        for branch in branches:
            packages = Application.objects.filter(
                status='Selected',
                student__branch=branch['branch']
            ).values_list('job__company__package', flat=True)
            
            if packages:
                branch['avg_package'] = float(sum(packages) / len(packages))
                branch['max_package'] = float(max(packages))
                branch['min_package'] = float(min(packages))
            else:
                branch['avg_package'] = 0
                branch['max_package'] = 0
                branch['min_package'] = 0
            
            branch['placement_rate'] = round(
                (branch['placed'] / branch['total_students'] * 100) if branch['total_students'] > 0 else 0, 2
            )
        
        report_data = branches
        export_headers = ['Branch', 'Total Students', 'Placed', 'Placement Rate %', 'Avg CGPA', 'Avg Package (LPA)', 'Max Package (LPA)', 'Min Package (LPA)']
        
    elif report_type == 'company':
        # Company-wise placement report
        companies = Company.objects.annotate(
            total_selected=Count('jobs__applications', filter=Q(jobs__applications__status='Selected')),
            unique_students=Count('jobs__applications__student', filter=Q(jobs__applications__status='Selected'), distinct=True)
        ).filter(total_selected__gt=0).order_by('-total_selected')
        
        report_data = companies
        export_headers = ['Company', 'Package (LPA)', 'Total Selected', 'Unique Students']
        
    else:  # student-wise
        # Student-wise placement details
        placed_apps = Application.objects.filter(
            status='Selected'
        ).select_related('student', 'job__company')
        
        report_data = []
        for app in placed_apps:
            report_data.append({
                'student': app.student,
                'company': app.job.company,
                'applied_date': app.applied_date,
            })
        
        export_headers = ['Name', 'Roll Number', 'Branch', 'Company', 'Package (LPA)', 'Applied Date']
    
    # Export functionality
    export_format = request.GET.get('export', '')
    if export_format:
        rows = []
        if report_type == 'branch':
            for branch in report_data:
                rows.append([
                    branch['branch'],
                    branch['total_students'],
                    branch['placed'],
                    branch['placement_rate'],
                    float(branch['avg_cgpa']) if branch['avg_cgpa'] else 0,
                    branch['avg_package'],
                    branch['max_package'],
                    branch['min_package'],
                ])
        elif report_type == 'company':
            for company in report_data:
                rows.append([
                    company.company_name,
                    float(company.package),
                    company.total_selected,
                    company.unique_students,
                ])
        else:  # student-wise
            for item in report_data:
                rows.append([
                    item['student'].user.get_full_name(),
                    item['student'].roll_number,
                    item['student'].branch,
                    item['company'].company_name,
                    float(item['company'].package),
                    item['applied_date'].strftime('%Y-%m-%d %H:%M'),
                ])
        
        if export_format == 'csv':
            return export_to_csv(f'placement_report_{report_type}', export_headers, rows)
        elif export_format == 'excel':
            return export_to_excel(f'placement_report_{report_type}', export_headers, rows, 'Placements')
    
    # Pagination
    paginator = Paginator(report_data, 20)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    
    context = {
        'report_type': report_type,
        'page_obj': page_obj,
    }
    return render(request, 'admin_panel/reports/placement_report.html', context)

@admin_required
def application_report(request):
    """
    Application status report with detailed filtering.
    """
    status_filter = request.GET.get('status', 'all')
    company_filter = request.GET.get('company', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    applications = Application.objects.select_related('student', 'job__company')
    
    if status_filter != 'all':
        applications = applications.filter(status=status_filter)
    
    if company_filter:
        applications = applications.filter(job__company_id=int(company_filter))
    
    if date_from:
        applications = applications.filter(applied_date__gte=date_from)
    
    if date_to:
        applications = applications.filter(applied_date__lte=date_to)
    
    # Status breakdown
    status_breakdown = {}
    for status in ['Applied', 'Shortlisted', 'Interview Scheduled', 'Selected', 'Rejected']:
        status_breakdown[status] = Application.objects.filter(status=status).count()
    
    # Export functionality
    export_format = request.GET.get('export', '')
    if export_format:
        headers = ['Student Name', 'Roll Number', 'Company', 'Job Title', 'Status', 'Applied Date', 'Last Updated']
        rows = []
        for app in applications:
            rows.append([
                app.student.user.get_full_name(),
                app.student.roll_number,
                app.job.company.company_name,
                app.job.job_title,
                app.status,
                app.applied_date.strftime('%Y-%m-%d'),
                app.updated_at.strftime('%Y-%m-%d'),
            ])
        
        if export_format == 'csv':
            return export_to_csv('application_report', headers, rows)
        elif export_format == 'excel':
            return export_to_excel('application_report', headers, rows, 'Applications')
    
    # Pagination
    paginator = Paginator(applications, 25)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    
    companies = Company.objects.all().order_by('company_name')
    statuses = ['Applied', 'Shortlisted', 'Interview Scheduled', 'Selected', 'Rejected']
    
    context = {
        'page_obj': page_obj,
        'status_filter': status_filter,
        'company_filter': company_filter,
        'date_from': date_from,
        'date_to': date_to,
        'companies': companies,
        'statuses': statuses,
        'status_breakdown': status_breakdown,
    }
    return render(request, 'admin_panel/reports/application_report.html', context)

@admin_required
def interview_report(request):
    """
    Interview report with scheduling and results.
    """
    interview_type_filter = request.GET.get('type', 'all')
    company_filter = request.GET.get('company', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    interviews = Interview.objects.select_related('application__student', 'application__job__company')
    
    if interview_type_filter != 'all':
        interviews = interviews.filter(interview_type=interview_type_filter)
    
    if company_filter:
        interviews = interviews.filter(application__job__company_id=int(company_filter))
    
    today = datetime.now().date()
    
    # Separate upcoming and completed interviews
    if request.GET.get('show') == 'completed':
        interviews = interviews.filter(date__lt=today).order_by('-date')
        show_type = 'completed'
    else:
        interviews = interviews.filter(date__gte=today).order_by('date')
        show_type = 'upcoming'
    
    if date_from:
        interviews = interviews.filter(date__gte=date_from)
    
    if date_to:
        interviews = interviews.filter(date__lte=date_to)
    
    # Export functionality
    export_format = request.GET.get('export', '')
    if export_format:
        headers = ['Student Name', 'Roll Number', 'Company', 'Interview Type', 'Date', 'Time', 'Venue', 'Status']
        rows = []
        today = datetime.now().date()
        for interview in interviews:
            status = 'Completed' if interview.date < today else 'Upcoming'
            rows.append([
                interview.application.student.user.get_full_name(),
                interview.application.student.roll_number,
                interview.application.job.company.company_name,
                interview.interview_type,
                interview.date.strftime('%Y-%m-%d'),
                interview.time.strftime('%H:%M'),
                interview.venue,
                status,
            ])
        
        if export_format == 'csv':
            return export_to_csv('interview_report', headers, rows)
        elif export_format == 'excel':
            return export_to_excel('interview_report', headers, rows, 'Interviews')
    
    # Pagination
    paginator = Paginator(interviews, 20)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    
    companies = Company.objects.all().order_by('company_name')
    interview_types = ['Technical', 'HR', 'Group Discussion', 'Final Round']
    
    # Statistics
    upcoming_count = Interview.objects.filter(date__gte=today).count()
    completed_count = Interview.objects.filter(date__lt=today).count()
    
    context = {
        'page_obj': page_obj,
        'interview_type_filter': interview_type_filter,
        'company_filter': company_filter,
        'date_from': date_from,
        'date_to': date_to,
        'show_type': show_type,
        'companies': companies,
        'interview_types': interview_types,
        'upcoming_count': upcoming_count,
        'completed_count': completed_count,
    }
    return render(request, 'admin_panel/reports/interview_report.html', context)


# ============================================================================
# PLACEMENT RESULT VIEWS (ADMIN)
# ============================================================================

@admin_required
@require_http_methods(["GET", "POST"])
@admin_required
@require_http_methods(["POST"])
def admin_revoke_result(request, placement_result_id):
    """
    Revoke a published placement result.
    
    Features:
    - Logs revocation with reason
    - Records revoked_by (current admin)
    - Records revoked_date
    - Generates notification for student
    - Maintains audit trail
    """
    from datetime import datetime
    
    placement_result = get_object_or_404(PlacementResult, id=placement_result_id)
    
    # Verify result is published
    if not placement_result.is_published:
        messages.error(request, 'Only published results can be revoked.')
        return redirect('admin_applications')
    
    # Check for revocation reason in POST data
    revocation_reason = request.POST.get('revocation_reason', '').strip()
    
    if not revocation_reason or len(revocation_reason) < 10:
        messages.error(request, 'Please provide a detailed reason for revocation (at least 10 characters).')
        return redirect('admin_applications')
    
    # Update placement result
    placement_result.is_revoked = True
    placement_result.revoked_by = request.user
    placement_result.revoked_date = datetime.now()
    placement_result.revocation_reason = revocation_reason
    placement_result.save()
    
    # Update application status to previous state (Shortlisted)
    application = placement_result.application
    application.status = 'Shortlisted'
    application.save()
    
    # Create revocation notification for student
    notification = Notification.objects.create(
        student=placement_result.student,
        notification_type='result_revoked',
        title='Placement Result Withdrawn',
        message=f"Your previously published placement result for {placement_result.company.company_name} "
                f"has been withdrawn by the Placement Cell. Please check the latest application status. "
                f"Reason: {revocation_reason}",
        application=application,
        placement_result=placement_result
    )
    
    messages.success(
        request,
        f'Placement result revoked for {placement_result.student.user.get_full_name()}. '
        f'Student has been notified.'
    )
    return redirect('admin_applications')


# ============================================================================
# NOTIFICATION VIEWS (STUDENT)
# ============================================================================

@login_required(login_url='login')
def student_notifications(request):
    """
    Display all notifications for the logged-in student.
    """
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        return redirect('student_profile_create')
    
    notifications = student.notifications.all().order_by('-created_at')
    
    # Pagination
    paginator = Paginator(notifications, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Mark all as read when viewing notifications page
    student.notifications.filter(is_read=False).update(is_read=True)
    
    context = {
        'page_obj': page_obj,
        'total_notifications': notifications.count(),
        'unread_count': 0,  # Already marked as read above
    }
    return render(request, 'student/notifications.html', context)


@login_required(login_url='login')
def mark_notification_read(request, notification_id):
    """
    Mark a specific notification as read and redirect back.
    """
    notification = get_object_or_404(Notification, id=notification_id)
    
    # Verify student owns this notification
    if notification.student.user != request.user:
        messages.error(request, 'Unauthorized access.')
        return redirect('home')
    
    notification.is_read = True
    notification.save()
    
    return redirect('student_notifications')


@login_required(login_url='login')
def get_unread_notifications_count(request):
    """
    AJAX endpoint to get unread notifications count for navbar.
    Returns JSON response.
    """
    try:
        student = request.user.student_profile
        unread_count = student.notifications.filter(is_read=False).count()
        return HttpResponse(json.dumps({'unread_count': unread_count}), content_type='application/json')
    except Student.DoesNotExist:
        return HttpResponse(json.dumps({'unread_count': 0}), content_type='application/json')


# ============================================================================
# PLACEMENT RESULTS DISPLAY (STUDENT)
# ============================================================================

@login_required(login_url='login')
def student_placement_results(request):
    """
    Display all placement results for the logged-in student.
    """
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        return redirect('student_profile_create')
    
    placement_results = student.placement_results.all().order_by('-selection_date')
    selected_results = placement_results.filter(status='Selected')
    
    context = {
        'placement_results': placement_results,
        'selected_results': selected_results,
        'has_selected': selected_results.exists(),
        'total_offers': placement_results.count(),
    }
    return render(request, 'student/placement_results.html', context)


# ============================================================================
# HELPER FUNCTION: HANDLE APPLICATION STATUS TRANSITIONS
# ============================================================================

def handle_application_status_transition(application, new_status, old_status, admin_user):
    """
    Handle all business logic when application status changes.
    Creates PlacementResult records, generates notifications, and maintains data consistency.
    
    Parameters:
    - application: Application object (already has new_status saved to database)
    - new_status: The new status that was just set
    - old_status: The previous status before the change
    - admin_user: The admin user making the change
    
    Status Transitions:
    - Applied → Shortlisted: Generate notification
    - Shortlisted → Interview Scheduled: Generate notification  
    - Interview Scheduled → Selected: Create PlacementResult, generate notification
    - Interview Scheduled → Rejected: Create PlacementResult, generate notification
    - Waiting List → Selected: Create PlacementResult, generate notification
    - Any status → Rejected: Create PlacementResult if not exists
    - Any status → Selected: Create PlacementResult if not exists
    """
    from datetime import date
    
    # Prevent duplicate notifications - only process if status actually changed
    if old_status == new_status:
        return False
    
    # ========================================================================
    # HANDLE FINAL STATUSES (Selected/Rejected) - CREATE PLACEMENT RESULT
    # ========================================================================
    
    if new_status in ['Selected', 'Rejected']:
        # Check if PlacementResult already exists and is published
        try:
            existing_result = application.placement_result
            if existing_result and existing_result.is_published and not existing_result.is_revoked:
                # Result already published - don't create duplicate
                return True
        except PlacementResult.DoesNotExist:
            pass
        
        # Create or update PlacementResult
        placement_result, created = PlacementResult.objects.get_or_create(
            application=application,
            defaults={
                'student': application.student,
                'company': application.job.company,
                'job': application.job,
                'status': new_status,
                'package_offered': application.job.company.package if new_status == 'Selected' else None,
                'selection_date': date.today(),
                'is_published': True,
                'published_by': admin_user,
                'is_revoked': False,
            }
        )
        
        # If result already existed but wasn't published, update it
        if not created and not placement_result.is_published:
            placement_result.status = new_status
            placement_result.selection_date = date.today()
            placement_result.is_published = True
            placement_result.published_by = admin_user
            if new_status == 'Selected':
                placement_result.package_offered = application.job.company.package
            placement_result.save()
        
        # Generate appropriate notification
        if new_status == 'Selected':
            create_notification(
                student=application.student,
                notification_type='selected',
                title='Placement Success',
                message=f"Congratulations! You have been selected by {application.job.company.company_name}. "
                        f"Package Offered: Rs. {placement_result.package_offered} LPA.",
                application=application,
                placement_result=placement_result
            )
        else:  # Rejected
            create_notification(
                student=application.student,
                notification_type='rejected',
                title='Application Status Updated',
                message=f"We regret to inform you that your application for {application.job.job_title} at {application.job.company.company_name} was not selected.",
                application=application,
                placement_result=placement_result
            )
        
        return True
    
    # ========================================================================
    # HANDLE INTERMEDIATE STATUSES - GENERATE NOTIFICATIONS ONLY
    # ========================================================================
    
    if new_status == 'Shortlisted' and old_status != 'Shortlisted':
        create_notification(
            student=application.student,
            notification_type='shortlisted',
            title='Application Shortlisted',
            message=f"Congratulations! You have been shortlisted for {application.job.job_title} at {application.job.company.company_name}. Please wait for further updates.",
            application=application
        )
        return True
    
    elif new_status == 'Interview Scheduled' and old_status != 'Interview Scheduled':
        create_notification(
            student=application.student,
            notification_type='interview_scheduled',
            title='Interview Scheduled',
            message=f"Your interview for {application.job.job_title} at {application.job.company.company_name} has been scheduled. Please check your email for details.",
            application=application
        )
        return True
    
    elif new_status == 'Interview Completed' and old_status != 'Interview Completed':
        create_notification(
            student=application.student,
            notification_type='interview_completed',
            title='Interview Update',
            message=f"Your interview for {application.job.job_title} at {application.job.company.company_name} has been completed. Please wait for the final result.",
            application=application
        )
        return True
    
    elif new_status == 'Waiting List' and old_status != 'Waiting List':
        create_notification(
            student=application.student,
            notification_type='waiting_list',
            title='Application Status Updated',
            message=f"Your application for {application.job.job_title} at {application.job.company.company_name} has been placed on the waiting list.",
            application=application
        )
        return True
    
    return True


# ============================================================================
# HELPER FUNCTION: CREATE NOTIFICATION
# ============================================================================

def create_notification(student, notification_type, title, message, application=None, job=None, placement_result=None):
    """
    Helper function to create a notification for a student.
    """
    notification = Notification.objects.create(
        student=student,
        notification_type=notification_type,
        title=title,
        message=message,
        application=application,
        job=job,
        placement_result=placement_result
    )
    return notification
